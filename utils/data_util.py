import json
import os
import re
import time
import openpyxl
import requests
from loguru import logger
from openpyxl.styles import Alignment, Font, PatternFill
from retry import retry


def norm_str(str):
    new_str = re.sub(r"|[\\/:*?\"<>| ]+", "", str).replace('\n', '').replace('\r', '')
    return new_str

def norm_text(text):
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    text = ILLEGAL_CHARACTERS_RE.sub(r'', text)
    return text


def timestamp_to_str(timestamp):
    timestamp = float(timestamp)
    # 同时兼容秒级和毫秒级时间戳
    if timestamp > 100000000000:
        timestamp /= 1000
    time_local = time.localtime(timestamp)
    dt = time.strftime("%Y-%m-%d %H:%M:%S", time_local)
    return dt

def extract_image_urls(images):
    """兼容抖音图片字段的新旧数据结构，统一返回 URL 列表。"""
    if not isinstance(images, list):
        return []

    image_urls = []
    for image in images:
        # 旧接口可能直接返回 URL 字符串
        if isinstance(image, str) and image:
            image_urls.append(image)
            continue
        if not isinstance(image, dict):
            continue

        # 新接口在图片对象的 url_list 中提供多个等价地址
        url_list = image.get('url_list', [])
        if isinstance(url_list, list):
            image_url = next((url for url in url_list if isinstance(url, str) and url), None)
            if image_url:
                image_urls.append(image_url)

    return image_urls


def infer_work_type(data, images, video_addr):
    """根据作品内容判断类型，兼容抖音新增的 aweme_type。"""
    # 图集优先，避免图集自带的视频字段造成误判
    if data.get('aweme_type') == 68 or images:
        return '图集'
    if isinstance(video_addr, str) and video_addr:
        return '视频'
    return '未知'

def handle_work_info(data):
    sec_uid = data['author']['sec_uid']
    user_url = f'https://www.douyin.com/user/{sec_uid}'
    user_desc = data['author']['signature'] if 'signature' in data['author'] else '未知'
    following_count = data['author']['following_count'] if 'following_count' in data['author'] else '未知'
    follower_count = data['author']['follower_count'] if 'follower_count' in data['author'] else '未知'
    total_favorited = data['author']['total_favorited'] if 'total_favorited' in data['author'] else '未知'
    aweme_count = data['author']['aweme_count'] if 'aweme_count' in data['author'] else '未知'
    user_id = data['author']['unique_id'] if 'unique_id' in data['author'] else '未知'
    user_age = data['author']['user_age'] if 'user_age' in data['author'] else '未知'
    gender = data['author']['gender'] if 'gender' in data['author'] else '未知'
    if gender == 1:
        gender = '男'
    elif gender == 0:
        gender = '女'
    else:
        gender = '未知'
    try:
        ip_location = data['user']['ip_location']
    except:
        ip_location = '未知'
    aweme_id = data['aweme_id']
    nickname = data['author']['nickname']
    author_avatar = data['author']['avatar_thumb']['url_list'][0]
    video_cover = data['video']['cover']['url_list'][0]
    title = data['desc']
    desc = data['desc']
    admire_count = data['statistics']['admire_count'] if 'admire_count' in data['statistics'] else 0
    digg_count = data['statistics']['digg_count']
    commnet_count = data['statistics']['comment_count']
    collect_count = data['statistics']['collect_count']
    share_count = data['statistics']['share_count']
    video_addr = data['video']['play_addr']['url_list'][0]
    images = extract_image_urls(data.get('images', []))
    create_time = data['create_time']

    text_extra = data['text_extra'] if 'text_extra' in data else []
    text_extra = text_extra if text_extra else []
    topics = []
    for item in text_extra:
        hashtag_name = item['hashtag_name'] if 'hashtag_name' in item else ''
        if hashtag_name:
            topics.append(hashtag_name)

    work_type = infer_work_type(data, images, video_addr)

    return {
        'work_id': aweme_id,
        'work_url': f'https://www.douyin.com/video/{aweme_id}',
        'work_type': work_type,
        'title': title,
        'desc': desc,
        'admire_count': admire_count,
        'digg_count': digg_count,
        'comment_count': commnet_count,
        'collect_count': collect_count,
        'share_count': share_count,
        'video_addr': video_addr,
        'images': images,
        'topics': topics,
        'create_time': create_time,
        'video_cover': video_cover,
        'user_url': user_url,
        'user_id': user_id,
        'nickname': nickname,
        'author_avatar': author_avatar,
        'user_desc': user_desc,
        'following_count': following_count,
        'follower_count': follower_count,
        'total_favorited': total_favorited,
        'aweme_count': aweme_count,
        'user_age': user_age,
        'gender': gender,
        'ip_location': ip_location
    }


def save_to_xlsx(datas, file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '作品信息'
    headers = ['作品id', '作品url', '作品类型', '作品标题', '描述', 'admire数量', '点赞数量', '评论数量', '收藏数量', '分享数量', '视频地址url', '图片地址url列表', '标签', '上传时间', '视频封面url', '用户主页url', '用户id', '昵称', '头像url', '用户描述', '关注数量', '粉丝数量', '作品被赞和收藏数量', '作品数量', '用户年龄', '性别', 'ip归属地']
    fields = ['work_id', 'work_url', 'work_type', 'title', 'desc', 'admire_count', 'digg_count',
              'comment_count', 'collect_count', 'share_count', 'video_addr', 'images', 'topics',
              'create_time', 'video_cover', 'user_url', 'user_id', 'nickname', 'author_avatar',
              'user_desc', 'following_count', 'follower_count', 'total_favorited', 'aweme_count',
              'user_age', 'gender', 'ip_location']
    ws.append(headers)
    for data in datas:
        row = []
        for field in fields:
            value = data.get(field, '')
            if field == 'create_time' and value not in ('', None):
                value = timestamp_to_str(value)
            elif isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)
            elif isinstance(value, str):
                value = norm_text(value)
            # ID 作为文本保存，避免科学计数法和精度丢失
            if field in ('work_id', 'user_id'):
                value = str(value)
            row.append(value)
        ws.append(row)

    # 设置适合阅读和筛选的基础样式
    header_fill = PatternFill('solid', fgColor='1F4E78')
    for cell in ws[1]:
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 24

    column_widths = {
        'A': 22, 'B': 42, 'C': 10, 'D': 42, 'E': 42,
        'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 12,
        'K': 42, 'L': 42, 'M': 30, 'N': 20, 'O': 42,
        'P': 42, 'Q': 18, 'R': 18, 'S': 42, 'T': 42,
        'U': 14, 'V': 14, 'W': 20, 'X': 14, 'Y': 12,
        'Z': 10, 'AA': 14,
    }
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top')
        for column in ('D', 'E', 'M', 'T'):
            ws[f'{column}{row[0].row}'].alignment = Alignment(vertical='top', wrap_text=True)
        ws.row_dimensions[row[0].row].height = 60
    for column in ('A', 'Q'):
        for cell in ws[column][1:]:
            cell.number_format = '@'
            cell.quotePrefix = True

    wb.save(file_path)
    logger.info(f'数据保存至 {file_path}')

def download_media(path, name, url, type):
    if type == 'image':
        content = requests.get(url).content
        with open(path + '/' + name + '.jpg', mode="wb") as f:
            f.write(content)
    elif type == 'video':
        res = requests.get(url, stream=True)
        size = 0
        chunk_size = 1024 * 1024
        with open(path + '/' + name + '.mp4', mode="wb") as f:
            for data in res.iter_content(chunk_size=chunk_size):
                f.write(data)
                size += len(data)


def save_wrok_detail(work, path):
    with open(f'{path}/detail.txt', mode="w", encoding="utf-8") as f:
        # 逐行输出到txt里
        f.write(f"作品id: {work['work_id']}\n")
        f.write(f"作品url: {work['work_url']}\n")
        f.write(f"作品类型: {work['work_type']}\n")
        f.write(f"作品标题: {work['title']}\n")
        f.write(f"描述: {work['desc']}\n")
        f.write(f"admire数量: {work['admire_count']}\n")
        f.write(f"点赞数量: {work['digg_count']}\n")
        f.write(f"评论数量: {work['comment_count']}\n")
        f.write(f"收藏数量: {work['collect_count']}\n")
        f.write(f"分享数量: {work['share_count']}\n")
        f.write(f"视频地址url: {work['video_addr']}\n")
        f.write(f"图片地址url列表: {', '.join(work['images'])}\n")
        f.write(f"标签: {', '.join(work['topics'])}\n")
        f.write(f"上传时间: {timestamp_to_str(work['create_time'])}\n")
        f.write(f"视频封面url: {work['video_cover']}\n")
        f.write(f"用户主页url: {work['user_url']}\n")
        f.write(f"用户id: {work['user_id']}\n")
        f.write(f"昵称: {work['nickname']}\n")
        f.write(f"头像url: {work['author_avatar']}\n")
        f.write(f"用户描述: {work['user_desc']}\n")
        f.write(f"关注数量: {work['following_count']}\n")
        f.write(f"粉丝数量: {work['follower_count']}\n")
        f.write(f"作品被赞和收藏数量: {work['total_favorited']}\n")
        f.write(f"作品数量: {work['aweme_count']}\n")
        f.write(f"用户年龄: {work['user_age']}\n")
        f.write(f"用户性别: {work['gender']}\n")
        f.write(f"ip归属地: {work['ip_location']}\n")


@retry(tries=3, delay=1)
def download_work(work_info, path, save_choice):
    work_id = work_info['work_id']
    user_id = work_info['user_id']
    title = work_info['title']
    title = norm_str(title)[:40]
    nickname = work_info['nickname']
    nickname = norm_str(nickname)[:20]
    if title.strip() == '':
        title = f'无标题'
    save_path = f'{path}/{nickname}_{user_id}/{title}_{work_id}'
    check_and_create_path(save_path)
    with open(f'{save_path}/info.json', mode='w', encoding='utf-8') as f:
        f.write(json.dumps(work_info) + '\n')
    work_type = work_info['work_type']
    save_wrok_detail(work_info, save_path)
    if work_type == '图集' and save_choice in ['media', 'media-image', 'all']:
        for img_index, img_url in enumerate(work_info['images']):
            download_media(save_path, f'image_{img_index}', img_url, 'image')
    elif work_type == '视频' and save_choice in ['media', 'media-video', 'all']:
        download_media(save_path, 'cover', work_info['video_cover'], 'image')
        download_media(save_path, 'video', work_info['video_addr'], 'video')
    logger.info(f'作品 {work_info["work_id"]} 下载完成，保存路径: {save_path}')
    return save_path



def check_and_create_path(path):
    if not os.path.exists(path):
        os.makedirs(path)
